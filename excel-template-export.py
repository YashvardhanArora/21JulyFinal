#!/usr/bin/env python3
"""
Excel Template Export with Full Formatting Preservation
Uses openpyxl to preserve all original Excel formatting, colors, borders, and dropdowns
"""

import sys
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def export_with_template(complaints_data, year=None):
    """
    Export complaints data using the original Excel template with preserved formatting
    """
    try:
        # Parse complaints data
        complaints = json.loads(complaints_data)
        
        # Filter by year if specified
        if year:
            filtered_complaints = []
            for complaint in complaints:
                if complaint.get('date'):
                    try:
                        complaint_year = datetime.fromisoformat(complaint['date'].replace('Z', '+00:00')).year
                        if str(complaint_year) == str(year):
                            filtered_complaints.append(complaint)
                    except:
                        pass
            complaints = filtered_complaints
        
        # Template file path
        template_path = Path('./attached_assets/2025final_1753735024170.xlsx')
        
        if not template_path.exists():
            print(json.dumps({"error": "Template file not found"}))
            return
        
        # Create a copy of the template to preserve original
        output_path = Path(f'/tmp/export-{year or "all"}-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        shutil.copy2(template_path, output_path)
        
        # Load the copied template workbook
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Clear existing data rows (keep row 1 as headers)
        max_row = worksheet.max_row
        if max_row > 1:
            worksheet.delete_rows(2, max_row)
        
        # Add complaint data starting from row 2
        for idx, complaint in enumerate(complaints, start=2):
            # Map complaint data to Excel columns (A-AD, 30 columns)
            row_data = [
                complaint.get('yearlySequenceNumber', idx-1),  # A: S.no.
                complaint.get('complaintSource', ''),          # B: Complaint Source
                complaint.get('placeOfSupply', ''),            # C: Place of Supply
                '',                                            # D: Complaint Receiving Location
                complaint.get('date', ''),                     # E: Month/Date
                complaint.get('depoPartyName', ''),            # F: Depo/Party Name
                complaint.get('email', ''),                    # G: Email
                complaint.get('contactNumber', ''),            # H: Contact Number
                complaint.get('invoiceNo', ''),                # I: Invoice No.
                complaint.get('date', ''),                     # J: Invoice Date
                complaint.get('lrNumber', ''),                 # K: LR Number
                complaint.get('transporterName', ''),          # L: Transporter Name
                complaint.get('transporterNumber', ''),        # M: Transporter Number
                complaint.get('complaintType', ''),            # N: Complaint Type
                complaint.get('voc', ''),                      # O: VOC
                complaint.get('salePersonName', ''),           # P: Sale Person Name
                complaint.get('productName', ''),              # Q: Product Name
                complaint.get('areaOfConcern', ''),            # R: Area of Concern
                complaint.get('subCategory', ''),              # S: Sub Category
                '',                                            # T: Action Taken
                '',                                            # U: Credit Date
                '',                                            # V: Credit Note Number
                '',                                            # W: Credit Amount
                '',                                            # X: Person Responsible
                '',                                            # Y: Root Cause/Action Plan
                complaint.get('createdAt', ''),                # Z: Complaint Creation
                '',                                            # AA: Date of Resolution
                '',                                            # AB: Date of Closure
                complaint.get('status', ''),                   # AC: Final Status
                ''                                             # AD: Days to Resolve
            ]
            
            # Write data to worksheet
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=idx, column=col_idx)
                cell.value = value if value != '' else '-'
        
        # Save the workbook with preserved formatting
        workbook.save(output_path)
        workbook.close()
        
        # Return the file path for download
        print(json.dumps({
            "success": True,
            "filePath": str(output_path),
            "count": len(complaints)
        }))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Complaints data required"}))
        sys.exit(1)
    
    complaints_data = sys.argv[1]
    year = sys.argv[2] if len(sys.argv) > 2 else None
    
    export_with_template(complaints_data, year)