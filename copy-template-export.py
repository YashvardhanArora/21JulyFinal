#!/usr/bin/env python3
"""
Direct Excel Template Copy with Minimal Changes
Preserves all original formatting by making minimal edits to the template
"""

import sys
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def export_with_direct_copy(complaints_data, year=None):
    """
    Export complaints by directly copying template and making minimal data changes
    """
    try:
        # Parse complaints data
        complaints = json.loads(complaints_data)
        
        # Filter by year if specified
        if year:
            filtered_complaints = []
            for complaint in complaints:
                if complaint.get('date'):
                    try:
                        complaint_year = datetime.fromisoformat(complaint['date'].replace('Z', '+00:00')).year
                        if str(complaint_year) == str(year):
                            filtered_complaints.append(complaint)
                    except:
                        pass
            complaints = filtered_complaints
        
        # Template file path
        template_path = Path('./attached_assets/2025final_1753735024170.xlsx')
        
        if not template_path.exists():
            print(json.dumps({"error": "Template file not found"}))
            return
        
        # Create output path
        output_path = Path(f'/tmp/formatted-export-{year or "all"}-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        # Copy the template file directly to preserve ALL formatting
        shutil.copy2(template_path, output_path)
        
        # Only modify data cells, keep ALL formatting intact
        workbook = load_workbook(output_path, keep_vba=True)
        worksheet = workbook.active
        
        # Get the range of existing data (starting from row 2, assuming row 1 is headers)
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Clear only cell VALUES from row 2 onwards, preserve all formatting
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                # Clear value but keep all formatting (font, fill, border, etc.)
                cell.value = None
        
        # Fill in complaint data starting from row 2
        for idx, complaint in enumerate(complaints):
            row_num = idx + 2  # Starting from row 2 (row 1 is headers)
            
            # Only set values, don't touch any formatting
            worksheet.cell(row=row_num, column=1).value = complaint.get('yearlySequenceNumber', idx + 1)
            worksheet.cell(row=row_num, column=2).value = complaint.get('complaintSource', '') or '-'
            worksheet.cell(row=row_num, column=3).value = complaint.get('placeOfSupply', '') or '-'
            worksheet.cell(row=row_num, column=4).value = '-'  # Complaint Receiving Location
            worksheet.cell(row=row_num, column=5).value = complaint.get('date', '') or '-'
            worksheet.cell(row=row_num, column=6).value = complaint.get('depoPartyName', '') or '-'
            worksheet.cell(row=row_num, column=7).value = complaint.get('email', '') or '-'
            worksheet.cell(row=row_num, column=8).value = complaint.get('contactNumber', '') or '-'
            worksheet.cell(row=row_num, column=9).value = complaint.get('invoiceNo', '') or '-'
            worksheet.cell(row=row_num, column=10).value = complaint.get('date', '') or '-'  # Invoice Date
            worksheet.cell(row=row_num, column=11).value = complaint.get('lrNumber', '') or '-'
            worksheet.cell(row=row_num, column=12).value = complaint.get('transporterName', '') or '-'
            worksheet.cell(row=row_num, column=13).value = complaint.get('transporterNumber', '') or '-'
            worksheet.cell(row=row_num, column=14).value = complaint.get('complaintType', '') or '-'
            worksheet.cell(row=row_num, column=15).value = complaint.get('voc', '') or '-'
            worksheet.cell(row=row_num, column=16).value = complaint.get('salePersonName', '') or '-'
            worksheet.cell(row=row_num, column=17).value = complaint.get('productName', '') or '-'
            worksheet.cell(row=row_num, column=18).value = complaint.get('areaOfConcern', '') or '-'
            worksheet.cell(row=row_num, column=19).value = complaint.get('subCategory', '') or '-'
            worksheet.cell(row=row_num, column=20).value = '-'  # Action Taken
            worksheet.cell(row=row_num, column=21).value = '-'  # Credit Date
            worksheet.cell(row=row_num, column=22).value = '-'  # Credit Note Number
            worksheet.cell(row=row_num, column=23).value = '-'  # Credit Amount
            worksheet.cell(row=row_num, column=24).value = '-'  # Person Responsible
            worksheet.cell(row=row_num, column=25).value = '-'  # Root Cause/Action Plan
            worksheet.cell(row=row_num, column=26).value = complaint.get('createdAt', '') or '-'
            worksheet.cell(row=row_num, column=27).value = '-'  # Date of Resolution
            worksheet.cell(row=row_num, column=28).value = '-'  # Date of Closure
            worksheet.cell(row=row_num, column=29).value = complaint.get('status', '') or '-'
            worksheet.cell(row=row_num, column=30).value = '-'  # Days to Resolve
        
        # Save with all original formatting preserved
        workbook.save(output_path)
        workbook.close()
        
        # Return the file path for download
        print(json.dumps({
            "success": True,
            "filePath": str(output_path),
            "count": len(complaints)
        }))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Complaints data required"}))
        sys.exit(1)
    
    complaints_data = sys.argv[1]
    year = sys.argv[2] if len(sys.argv) > 2 else None
    
    export_with_direct_copy(complaints_data, year)